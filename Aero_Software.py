from gpuinfo import GPUInfo
from psutil import virtual_memory
from os import environ,path
from math import ceil
from psutil import net_if_stats,disk_partitions
from wmi import WMI
from time import strftime,localtime
from openpyxl import Workbook
from openpyxl import load_workbook 
from ssd import is_ssd
from sys import exit


filepath = "dane.txt"  # ścieżka relative
try:
    with open(filepath, "r") as f:
        lokalizacja = f.read()
        f.close()
except FileNotFoundError:
    print("Auto path file not found.")
    lokalizacja='podzespoly.xlsx'

if path.exists(lokalizacja):
    print()
    print('Location - ',lokalizacja)
else:
    print()
    print("Directory does not exist. ",lokalizacja)
    lokalizacja='podzespoly.xlsx'
    print("Saves to local folder!!! ",lokalizacja)
  
  
def load_to_file():
    try:
        wb = load_workbook(lokalizacja) 
    except FileNotFoundError:
        wb = Workbook()
        wb.create_sheet()
        print("The new sheet has been created!!!!!")
        print()
        
    sheet = wb.active
    
    sheet.append((czas,manager,room_nb,section,nazwa_komputer,uzytkownik,board,procesor,ile_ram,grafika,karta_sieciowa,dysk_ssd))
    try:
        wb.save(lokalizacja)
        print()
        print('Success, data saved ')
        print()
    except:
        print("You do not have permission to write to this folder!!!!")
        print("Maybe someone will use this file")
        wybor= input('You whant try again? Y/N - ')
        print()
        while True:

            if wybor=='Y'or wybor=='y':
                load_to_file()
                break
            elif wybor=='N'or wybor=='n':
                print('Data has not been saved to the file!')
                exit(0)
            else:
                print("You mast write Y or N ")
                wybor= input('You whant try again? Y/N - ')
                print()
            

def wyswietl():
    print('1:Time',czas)
    print('2:Manager ',manager)
    print('3:Room number ',room_nb)
    print("4:Section ", section)
    print("5:Computer_name ",nazwa_komputer)
    print('6:User =',uzytkownik)
    print('7:Motherboard =',board )
    print('8:CPU: {0}'.format(procesor))
    print('9:RAM: {0}GB'.format(ile_ram))
    print('10:Graphics Card: {0}'.format(grafika))
    print('11:Network card= {0}'.format(karta_sieciowa))
    print("12:SSD data disck ", dysk_ssd)


computer = WMI()

#computer_info = computer.Win32_ComputerSystem()[0]

os_info = computer.Win32_OperatingSystem()[0]
czas=strftime("%Y-%m-%d %H:%M:%S", localtime())
nazwa_komputer=environ['COMPUTERNAME']

system_ram = float(os_info.TotalVisibleMemorySize) / 1048576  # KB to GB
ile_ram=ceil(system_ram)

proc_info = computer.Win32_Processor()[0]
procesor = proc_info.Name

gpu_info = computer.Win32_VideoController()[0]
grafika =gpu_info.Name

stats = net_if_stats()
#domyslna_karta = stats["Ethernet"]
predkosc=1000
#print(predkosc)
for isup in stats:
    st = stats[isup]
    stspeed=st.speed
    if predkosc < stspeed:
        predkosc = stspeed
predkosc = round(predkosc/1024)
if predkosc > 1.5:
    predkosc=10       
      
karta_sieciowa =predkosc

#karta_sieciowa = ceil(st.speed/1024)
#print(stats)
# siec= computer.Win32_NetworkAdapter()[0]
# siec_gotowa=siec.NetConnectionID
# print(siec_gotowa)


#print(plyta_glowna)
#os_name = os_info.Name.encode('utf-8').split(b'|')[0]
#os_version = ' '.join([os_info.Version, os_info.BuildNumber])
#print('OS Name: {0}'.format(os_name))
#print('OS Version: {0}'.format(os_version))
#print('RAM: {0} GB'.format(system_ram))
print()
manager=input("Manager - ")
room_nb=input("Room number - ")
section=input("Sectiont - ")
uzytkownik=input('User - ')
print()


dyski =disk_partitions(all=False)
dysk_ssd=0
## Sprawdzanie czy ssd
for i in range(len(dyski)):
  sciezka=(dyski[i].device)+'/'
  dysk_czy_ssd=(is_ssd(sciezka))
  if dysk_czy_ssd == True:
    dysk_ssd+=1
## Wykluczenie dysku systemowego    
  
if dysk_ssd != 0:
    dysk_ssd-=1


motherboard=computer.Win32_BaseBoard()[0]
board=motherboard.Manufacturer + motherboard.Product


load_to_file()
wyswietl()
input("Press Enter to continue...")



